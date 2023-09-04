from .errors import EmptySegmentError, ShortSegmentError
from .ui_mr_main_window import Ui_MrMainWindow
from PyQt6.QtWidgets import QMainWindow, QVBoxLayout, QFileDialog, QMessageBox
from PyQt6.QtCore import Qt, QSettings
from matplotlib.figure import Figure
from matplotlib.backends.backend_qtagg import FigureCanvas, NavigationToolbar2QT
from .sequence import Sequence
from os.path import dirname, exists
from openpyxl import Workbook, load_workbook
from .on_save_excel_dialog import OnSaveExcelDialog


class MrMainWindow(QMainWindow):
    def __init__(self, ds_mr_signal, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.__lines = []
        self.__interpolation_lines = []
        self.__raw_line = None

        self.__ds_mr_signal = ds_mr_signal
        self.__settings = QSettings()

        self.__ui = Ui_MrMainWindow()
        self.__ui.setupUi(self)

        self.mr_calculated = True

        self.__ui.save_excel_action.triggered.connect(self.__save_excel)
        self.__ui.close_action.triggered.connect(self.__close)

        self.__ru = self.__settings.value("ru_mnemonics", type=dict)

        self.__colors = self.__settings.value("colors", type=dict)

        figure = Figure(figsize=(5, 3))
        self.__canvas = FigureCanvas(figure)
        self.__ax = self.__canvas.figure.subplots()
        self.__ax.grid()
        self.__ax.set_xlabel(str(self.__ds_mr_signal.cols.etalon_pq))
        self.__ax.set_ylabel(str(self.__ds_mr_signal.cols.dut))

        layout = QVBoxLayout(self.__ui.centralwidget)
        layout.addWidget(self.__canvas)

        toolbar = NavigationToolbar2QT(self.__canvas, self)
        self.addToolBar(toolbar)
        toolbar.setAllowedAreas(Qt.ToolBarArea.AllToolBarAreas)

        sequence = Sequence(self.__settings.value("sequence", type=int))
        interpolate = self.__settings.value("grid/interpolate", type=bool)
        detector_type = self.__settings.value("grid/detector_type", type=str)
        detector_value = self.__settings.value("grid/detector_value", type=int)

        if self.__settings.value("grid/on", type=bool):
            sequence = Sequence(self.__settings.value("sequence", type=int))
            margin = self.__settings.value("grid/margin", type=float)
            grid = [float(item)
                    for item in self.__settings.value("grid/data", type=list)]
            if interpolate:
                self.__ds_mr_signal.add_mr_interpolated_handler(
                    self.__mr_interpolated_handler)
            else:
                self.__ds_mr_signal.add_mr_calculated_handler(
                    self.__mr_calculated_handler)
            if self.__settings.value("down_grid/on", type=bool):
                down_grid = [float(item)
                             for item in self.__settings.value("down_grid/data", type=list)]
                try:
                    self.__ds_mr_signal.calculate_parts_and_segments_by_grid(
                        sequence, margin, detector_type, detector_value, grid, down_grid, interpolate)
                except EmptySegmentError as err:
                    if err.type == "up":
                        x = "подъема"
                    elif err.type == "down":
                        x = "спуска"
                    QMessageBox.critical(
                        None, "Ошибка", (f"Сегмент {x} {str(err.grid_value)} {self.__ds_mr_signal.cols.etalon_pq.unit}"
                                         " слишком короткий. Проверьте сетку и повторите попытку"))
                    self.__ds_mr_signal.remove_mr_interpolated_handler(
                        self.__mr_interpolated_handler)
                    self.__ds_mr_signal.remove_mr_calculated_handler(
                        self.__mr_calculated_handler)
                    self.mr_calculated = False
                except ShortSegmentError as err:
                    if err.type == "up":
                        x = "подъема"
                    elif err.type == "down":
                        x = "спуска"
                    QMessageBox.critical(
                        None, "Ошибка", (f"Сегмент {x} {str(err.grid_value)} {self.__ds_mr_signal.cols.etalon_pq.unit}"
                                         " слишком короткий для текущего детектора. Проверьте детектор и повторите попытку"))
                    self.__ds_mr_signal.remove_mr_interpolated_handler(
                        self.__mr_interpolated_handler)
                    self.__ds_mr_signal.remove_mr_calculated_handler(
                        self.__mr_calculated_handler)
                    self.mr_calculated = False
            else:
                try:
                    self.__ds_mr_signal.calculate_parts_and_segments_by_grid(
                        sequence, margin, detector_type, detector_value, grid, down_grid=None, interpolate=interpolate)
                except EmptySegmentError as err:
                    if err.type == "up":
                        x = "подъема"
                    elif err.type == "down":
                        x = "спуска"
                    QMessageBox.critical(
                        None, "Ошибка", (f"Сегмент {x} {str(err.grid_value)} {self.__ds_mr_signal.cols.etalon_pq.unit}"
                                         " слишком короткий. Проверьте сетку и повторите попытку"))
                    self.__ds_mr_signal.remove_mr_interpolated_handler(
                        self.__mr_interpolated_handler)
                    self.__ds_mr_signal.remove_mr_calculated_handler(
                        self.__mr_calculated_handler)
                    self.mr_calculated = False
                except ShortSegmentError as err:
                    if err.type == "up":
                        x = "подъема"
                    elif err.type == "down":
                        x = "спуска"
                    QMessageBox.critical(
                        None, "Ошибка", (f"Сегмент {x} {str(err.grid_value)} {self.__ds_mr_signal.cols.etalon_pq.unit}"
                                         " слишком короткий для текущего детектора. Проверьте детектор и повторите попытку"))
                    self.__ds_mr_signal.remove_mr_interpolated_handler(
                        self.__mr_interpolated_handler)
                    self.__ds_mr_signal.remove_mr_calculated_handler(
                        self.__mr_calculated_handler)
                    self.mr_calculated = False
        else:
            self.__ds_mr_signal.add_parts_calculated_handler(
                self.__parts_calculated_handler)
            self.__ds_mr_signal.calculate_parts_by_extremum(sequence)

    def __parts_calculated_handler(self):
        self.__clear_lines()

        for part in self.__ds_mr_signal.parts:
            if part.type == "up":
                label = "подъем"
            elif part.type == "down":
                label = "спуск"
            line, = self.__ax.plot(part[str(self.__ds_mr_signal.cols.etalon_pq)],
                                   part[str(self.__ds_mr_signal.cols.dut)],
                                   label=label, color=self.__colors[part.type])
            self.__lines.append(line)

        self.__ax.legend()

    def __clear_lines(self):
        for line in self.__lines:
            self.__ax.lines.remove(line)
        self.__lines.clear()

    def __clear_interpolation_lines(self):
        for line in self.__interpolation_lines:
            self.__ax.lines.remove(line)
        self.__interpolation_lines.clear()

    def __clear_raw_line(self):
        if self.__raw_line:
            self.__ax.lines.remove(self.__raw_line)
        self.__raw_line = None

    def __mr_calculated_handler(self):
        self.__clear_lines()

        if (inds := self.__ds_mr_signal.get_up_mr_inds()) is not None:
            line, = self.__ax.plot(self.__ds_mr_signal.df.loc[inds, str(
                self.__ds_mr_signal.cols.etalon_pq)],
                self.__ds_mr_signal.df.loc[inds, str(
                    self.__ds_mr_signal.cols.dut)],
                label=self.__ru["up"], color=self.__colors["up"],
                marker=".")
            self.__lines.append(line)

        if (inds := self.__ds_mr_signal.get_down_mr_inds()) is not None:
            line, = self.__ax.plot(self.__ds_mr_signal.df.loc[inds, str(
                self.__ds_mr_signal.cols.etalon_pq)],
                self.__ds_mr_signal.df.loc[inds, str(
                    self.__ds_mr_signal.cols.dut)],
                label=self.__ru["down"], color=self.__colors["down"],
                marker=".")
            self.__lines.append(line)

        if self.__settings.value("show_raw_magnitude_response", type=bool):
            self.__plot_raw_line()

        self.__ax.legend()
        self.__canvas.draw_idle()

    def __mr_interpolated_handler(self):
        self.__clear_interpolation_lines()

        for part_type in self.__ds_mr_signal.interpolated_mr:
            line, = self.__ax.plot(
                self.__ds_mr_signal.interpolated_mr[part_type]["x"],
                self.__ds_mr_signal.interpolated_mr[part_type]["y"],
                label=self.__ru[part_type], color=self.__colors[part_type],
                marker=".")
            self.__interpolation_lines.append(line)

        if self.__settings.value("show_raw_magnitude_response", type=bool):
            self.__plot_raw_line()

        self.__ax.legend()
        self.__canvas.draw_idle()

    def __plot_raw_line(self, draw_idle=False):
        self.__clear_raw_line()

        raw_line, = self.__ax.plot(
            self.__ds_mr_signal.df[str(self.__ds_mr_signal.cols.etalon_pq)],
            self.__ds_mr_signal.df[str(self.__ds_mr_signal.cols.dut)],
            label=self.__ru["raw"],
            zorder=0,
            color=self.__colors["raw"])

        self.__raw_line = raw_line

    def __save_excel(self):
        file_path = self.__get_save_file_path()

        if not file_path:
            return

        enable_append_check_box = False
        if exists(file_path):
            enable_append_check_box = True

        res = OnSaveExcelDialog(
            enable_append_check_box=enable_append_check_box).exec()

        if res == 1:
            if exists(file_path) and self.__settings.value("output/excel/append", type=bool):
                wb = load_workbook(file_path)
            else:
                wb = Workbook()

            ws = wb.active
            start_cell = ws[self.__settings.value("output/excel/start_cell")]
            start_row = start_cell.row
            start_col_idx = start_cell.col_idx
            cur_row = start_row
            cur_col_idx = start_col_idx

            if self.__settings.value("grid/interpolate", type=bool):
                if "up" in self.__ds_mr_signal.interpolated_mr:
                    if self.__settings.value("output/excel/direction") == "horizontal":
                        ws.cell(cur_row, cur_col_idx, self.__ru["up"])
                        cur_row += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.etalon_pq))
                        cur_row += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.dut))
                        cur_row -= 1
                        cur_col_idx += 1
                        for i, val in enumerate(self.__ds_mr_signal.interpolated_mr["up"]["x"]):
                            ws.cell(cur_row, cur_col_idx, val)
                            cur_row += 1
                            ws.cell(cur_row, cur_col_idx,
                                    self.__ds_mr_signal.interpolated_mr["up"]["y"][i])
                            cur_col_idx += 1
                            cur_row -= 1
                        cur_row = start_row + 3
                        cur_col_idx = start_col_idx
                    elif self.__settings.value("output/excel/direction") == "vertical":
                        ws.cell(cur_row, cur_col_idx, self.__ru["up"])
                        cur_row += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.etalon_pq))
                        cur_col_idx += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.dut))
                        cur_col_idx -= 1
                        cur_row += 1
                        for i, val in enumerate(self.__ds_mr_signal.interpolated_mr["up"]["x"]):
                            ws.cell(cur_row, cur_col_idx, val)
                            cur_col_idx += 1
                            ws.cell(cur_row, cur_col_idx,
                                    self.__ds_mr_signal.interpolated_mr["up"]["y"][i])
                            cur_row += 1
                            cur_col_idx -= 1
                        cur_row = start_row
                        cur_col_idx = start_col_idx + 2

                if "down" in self.__ds_mr_signal.interpolated_mr:
                    write_x = True
                    if (("up" in self.__ds_mr_signal.interpolated_mr) and
                            self.__ds_mr_signal.interpolated_mr["up"]["x"] == self.__ds_mr_signal.interpolated_mr["down"]["x"]):
                        write_x = False
                    if self.__settings.value("output/excel/direction") == "horizontal":
                        ws.cell(cur_row, cur_col_idx, self.__ru["down"])
                        cur_row += 1
                        if write_x:
                            ws.cell(cur_row, cur_col_idx, str(
                                self.__ds_mr_signal.cols.etalon_pq))
                            cur_row += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.dut))
                        if write_x:
                            cur_row -= 1
                        cur_col_idx += 1
                        for i, val in enumerate(self.__ds_mr_signal.interpolated_mr["down"]["x"]):
                            if write_x:
                                ws.cell(cur_row, cur_col_idx, val)
                                cur_row += 1
                            ws.cell(cur_row, cur_col_idx,
                                    self.__ds_mr_signal.interpolated_mr["down"]["y"][i])
                            cur_col_idx += 1
                            if write_x:
                                cur_row -= 1
                        cur_row = start_row + 3
                        cur_col_idx = start_col_idx
                    elif self.__settings.value("output/excel/direction") == "vertical":
                        ws.cell(cur_row, cur_col_idx, self.__ru["down"])
                        cur_row += 1
                        if write_x:
                            ws.cell(cur_row, cur_col_idx, str(
                                self.__ds_mr_signal.cols.etalon_pq))
                            cur_col_idx += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.dut))
                        if write_x:
                            cur_col_idx -= 1
                        cur_row += 1
                        for i, val in enumerate(self.__ds_mr_signal.interpolated_mr["down"]["x"]):
                            if write_x:
                                ws.cell(cur_row, cur_col_idx, val)
                                cur_col_idx += 1
                            ws.cell(cur_row, cur_col_idx,
                                    self.__ds_mr_signal.interpolated_mr["down"]["y"][i])
                            cur_row += 1
                            if write_x:
                                cur_col_idx -= 1
                        cur_row = start_row
                        cur_col_idx = start_col_idx + 2
            else:
                if (inds := self.__ds_mr_signal.get_up_mr_inds()) is not None:
                    if self.__settings.value("output/excel/direction") == "horizontal":
                        ws.cell(cur_row, cur_col_idx, self.__ru["up"])
                        cur_row += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.etalon_pq))
                        cur_row += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.dut))
                        cur_row -= 1
                        cur_col_idx += 1
                        for i, row in self.__ds_mr_signal.df.loc[inds].sort_values(
                                by=str(self.__ds_mr_signal.cols.etalon_pq)).iterrows():
                            ws.cell(cur_row, cur_col_idx, row[str(
                                self.__ds_mr_signal.cols.etalon_pq)])
                            cur_row += 1
                            ws.cell(cur_row, cur_col_idx,
                                    row[str(self.__ds_mr_signal.cols.dut)])
                            cur_col_idx += 1
                            cur_row -= 1
                        cur_row = start_row + 3
                        cur_col_idx = start_col_idx
                    elif self.__settings.value("output/excel/direction") == "vertical":
                        ws.cell(cur_row, cur_col_idx, self.__ru["up"])
                        cur_row += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.etalon_pq))
                        cur_col_idx += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.dut))
                        cur_col_idx -= 1
                        cur_row += 1
                        for i, row in self.__ds_mr_signal.df.loc[inds].sort_values(
                                by=str(self.__ds_mr_signal.cols.etalon_pq)).iterrows():
                            ws.cell(cur_row, cur_col_idx, row[str(
                                self.__ds_mr_signal.cols.etalon_pq)])
                            cur_col_idx += 1
                            ws.cell(cur_row, cur_col_idx,
                                    row[str(self.__ds_mr_signal.cols.dut)])
                            cur_row += 1
                            cur_col_idx -= 1
                        cur_row = start_row
                        cur_col_idx = start_col_idx + 2

                if (inds := self.__ds_mr_signal.get_down_mr_inds()) is not None:
                    if self.__settings.value("output/excel/direction") == "horizontal":
                        ws.cell(cur_row, cur_col_idx, self.__ru["down"])
                        cur_row += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.etalon_pq))
                        cur_row += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.dut))
                        cur_row -= 1
                        cur_col_idx += 1
                        for i, row in self.__ds_mr_signal.df.loc[inds].sort_values(
                                by=str(self.__ds_mr_signal.cols.etalon_pq)).iterrows():
                            ws.cell(cur_row, cur_col_idx, row[str(
                                self.__ds_mr_signal.cols.etalon_pq)])
                            cur_row += 1
                            ws.cell(cur_row, cur_col_idx,
                                    row[str(self.__ds_mr_signal.cols.dut)])
                            cur_col_idx += 1
                            cur_row -= 1
                    elif self.__settings.value("output/excel/direction") == "vertical":
                        ws.cell(cur_row, cur_col_idx, self.__ru["down"])
                        cur_row += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.etalon_pq))
                        cur_col_idx += 1
                        ws.cell(cur_row, cur_col_idx, str(
                            self.__ds_mr_signal.cols.dut))
                        cur_col_idx -= 1
                        cur_row += 1
                        for i, row in self.__ds_mr_signal.df.loc[inds].sort_values(
                                by=str(self.__ds_mr_signal.cols.etalon_pq)).iterrows():
                            ws.cell(cur_row, cur_col_idx, row[str(
                                self.__ds_mr_signal.cols.etalon_pq)])
                            cur_col_idx += 1
                            ws.cell(cur_row, cur_col_idx,
                                    row[str(self.__ds_mr_signal.cols.dut)])
                            cur_row += 1
                            cur_col_idx -= 1

            try:
                wb.save(file_path)
            except PermissionError:
                QMessageBox.critical(
                    None, "Ошибка",
                    "Выбранный файл открыт в другой программе. Закройте его и повторите попытку")

    def __get_save_file_path(self):
        reports_dir = self.__settings.value("reports_dir")
        gtr_dir = self.__settings.value("gtr_dir")
        file = QFileDialog.getSaveFileName(
            self, "Сохранить txt-файл", reports_dir, "Файлы excel (*.xlsx)")
        file_path = file[0]
        dir_path = dirname(file_path)
        if dir_path == gtr_dir:
            self.__settings.setValue("use_same_gtr_and_reports_dir", True)
        else:
            self.__settings.setValue("use_same_gtr_and_reports_dir", False)
            self.__settings.setValue("reports_dir", dir_path)
        return file_path

    def __close(self):
        self.close()

    def closeEvent(self, a0):
        self.__ds_mr_signal.remove_parts_calculated_handler(
            self.__parts_calculated_handler)
        self.__ds_mr_signal.remove_mr_calculated_handler(
            self.__mr_calculated_handler)
        self.__ds_mr_signal.remove_mr_interpolated_handler(
            self.__mr_interpolated_handler)
        super().closeEvent(a0)
